import requests
import pandas as pd
import logging
import time
import sqlite3
import json
import signal
import sys
import os
import traceback
import threading
import csv
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Casey Staples
# 18/11/23

# globals cuz I'm getting lazy
interrupted = False
excluded_keywords = ['hentai', 'sex', 'nazi', 'sexual', 'daddy', 'mommy', 'furry', 'college', 'fuck', 'fucking', 'ass', 'dick', 'penis', 'cock',
                     'boob', 'tit', 'trump', 'obama', 'biden', 'anime', 'test', 'pack', 'dlc', 'soundtrack', 'demo']
game_data = []  # so I can try threading
data_lock = threading.Lock()  # Need to setup a lock for game_data


# Configure logging
# To turn off info change logging level to logging.CRITICAL
# to see debugs set to debug (you wont need to)
# To stop saving file, remove the file handler
# If you want to stop terminal output, remove console handler
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        # file handler
                        logging.StreamHandler()  # console handler
                    ])

# Wanted to see if I can get threading to work for fun


def save_to_db_interval(interval, stop_event):
    try:
        while not stop_event.is_set():
            # Perform the save operation
            with data_lock:
                write_data_to_sqlite(game_data)

            # Wait for the specified interval, checking each second if the stop event is set
            # Otherwise it will wait the full interval before checking again, causing the script to take longer to stop
            for _ in range(interval):
                if stop_event.is_set():
                    break
                time.sleep(1)

    except Exception as e:
        logging.critical(f"Error saving data to SQLite database: {e}")
        logging.critical("Full traceback:" + '\n' + traceback.format_exc())
    finally:
        logging.info("Saving thread cleanup completed and exiting.")


def write_data_to_sqlite(game_data, db='gameData.db'):
    # Create a DataFrame from the game data
    df_new_data = pd.DataFrame(game_data)

    # Establish a connection to the SQLite database
    conn = sqlite3.connect(db)

    try:
        # Create a table with the appropriate schema
        # This schema should match the keys in game_data if I can read
        # Could probably move this to check only once, but I'm going to keep it here for now
        conn.execute('''
            CREATE TABLE IF NOT EXISTS steam_data (
                Name TEXT,
                AppID INTEGER PRIMARY KEY,
                Short_Description TEXT,
                Review_Score INTEGER,
                Review_Score_Description TEXT,
                Total_Positive INTEGER,
                Total_Negative INTEGER,
                Total_Reviews INTEGER,
                Positive_Review_Percentage REAL,
                Review_Text TEXT,
                Playtime_Forever TEXT,
                Playtime_Last_Two_Weeks TEXT,
                Playtime_at_Review TEXT,
                Timestamp_Created TEXT,
                Timestamp_Updated TEXT,
                Voted_Up TEXT,
                Votes_Up TEXT,
                Genres TEXT
            )
        ''')

        # Append data to the table
        df_new_data.to_sql('steam_data', conn, if_exists='append', index=False)

        logging.info(f"Data successfully saved to {db}")

    except Exception as e:
        logging.critical(f"Failed to write data to SQLite database: {e}")

    finally:
        # Close the database connection
        conn.close()


def load_unwanted_games():
    try:
        with open('unwanted_apps.json', 'r') as f:
            unwanted_apps = json.load(f)
            logging.info("Loaded list of games without reviews.")
            return unwanted_apps
    except FileNotFoundError:
        unwanted_apps = []
        logging.warning("No existing list of unwanted games found.")
        return unwanted_apps


# This are game ids that are dlcs or have no reviews.
def save_unwanted_apps_list(unwanted_apps):
    try:
        filename = 'unwanted_apps.json'
        # Check if file exists and is not empty
        if os.path.exists(filename) and os.path.getsize(filename) > 0:
            with open(filename, 'r+') as file:
                # Load existing data
                existing_data = json.load(file)
                # Update data
                # Remove duplicates
                updated_data = list(set(existing_data + unwanted_apps))
                # Rewind file to the beginning
                file.seek(0)
                # Dump updated data
                json.dump(updated_data, file)
                # Truncate file to new size
                file.truncate()
        else:
            # If file doesn't exist or is empty, write new data
            with open(filename, 'w') as file:
                json.dump(unwanted_apps, file)

        logging.info(f"Saved unwanted games data to {filename}")

    except Exception as e:
        logging.critical(f"Error saving unwanted games data: {e}")

# functions for saving and loading the appid list, so we can resume from where we left off


def save_remaning_apps_list(appid_data, filename='remaining_games.json'):
    try:
        with open(filename, 'w') as file:
            json.dump(appid_data, file)
        logging.info(f"Saved AppID data to {filename}")
    except Exception as e:
        logging.critical(f"Error saving AppID data: {e}")


def load_remaning_app_list(filename='remaining_games.json'):
    try:
        if os.path.exists(filename):
            with open(filename, 'r') as file:
                logging.info(f"Loading AppID data from {filename}")
                return json.load(file)
    except Exception as e:
        logging.critical(f"Error loading AppID data: {e}")
    return None

# Auto adjust column widths after writing to Excel since it's going to be Yuge
# Deprecated


def auto_adjust_columns(filepath):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    # Word wrap for the description and review text columns
    wrap_columns = ['C', 'J']

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[col_letter].width = length

        if col_letter in wrap_columns:
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True)

    workbook.save(filepath)
    workbook.close()
    logging.info("Column widths adjusted.")

# Function to write data to Excel


def export_db_to_csv(db='gameData.db', csv_filename='steam_data.csv'):
    try:
        # Establish a connection to the SQLite database
        conn = sqlite3.connect(db)
        cursor = conn.cursor()

        # Query to select all data from the table
        query = "SELECT * FROM steam_data"
        cursor.execute(query)

        # Fetch all the data
        data = cursor.fetchall()

        # If there's no data, exit the function
        if not data:
            logging.info("No data available to write to CSV.")
            return

        # Column headers based on your data structure
        headers = ['Name', 'AppID', 'Short Description', 'Review Score', 'Review Score Description',
                   'Total Positive', 'Total Negative', 'Total Reviews', 'Positive Review Percentage',
                   'Review Text (First Review)', 'Playtime Forever (First Review)',
                   'Playtime Last Two Weeks (First Review)', 'Playtime at Review (First Review)',
                   'Timestamp Created (First Review)', 'Timestamp Updated (First Review)',
                   'Voted Up (First Review)', 'Votes Up (First Review)', 'Genres']

        # Write data to CSV file
        with open(csv_filename, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file, lineterminator='\n')
            writer.writerow(headers)  # Write the header
            writer.writerows(data)  # Write the data rows

        logging.info(f"Data successfully exported to {csv_filename}")

    except Exception as e:
        logging.critical(f"Failed to export data to CSV file: {e}")

    finally:
        # Close the database connection
        conn.close()

# Deprecated


def export_db_to_excel(db='gameData.db', excel_filename='steam_data.xlsx'):
    try:
        # Establish a connection to the SQLite database
        conn = sqlite3.connect(db)

        # Read the entire table into a DataFrame
        query = "SELECT * FROM steam_data"
        df = pd.read_sql(query, conn)

        # Write (or overwrite) the DataFrame to an Excel file
        df.to_excel(excel_filename, index=False)
        logging.info(f"Data successfully exported to {excel_filename}")
        auto_adjust_columns(excel_filename)

    except Exception as e:
        logging.critical(f"Failed to export data to Excel file: {e}")

    finally:
        # Close the database connection
        conn.close()


# check for crt+c in terminal to stop the script and write the data to excel
def signal_handler(sig, frame):
    global interrupted
    interrupted = True
    logging.info("Ctrl+C detected. Preparing to save data.")


def is_english(s):
    try:
        s.encode(encoding='utf-8').decode('ascii')
    except UnicodeDecodeError:
        return False
    else:
        return True


def contains_excluded_keywords(s, keywords):
    return any(keyword in s.lower() for keyword in keywords)


def get_steam_games():
    url = "https://api.steampowered.com/ISteamApps/GetAppList/v2/"
    response = requests.get(url)
    data = response.json()
    apps = data["applist"]["apps"]

    return [app for app in apps if app['name'].strip()
            and is_english(app['name'])
            and not contains_excluded_keywords(app['name'], excluded_keywords)]


def get_game_details(appid):
    url = f"http://store.steampowered.com/api/appdetails?appids={appid}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()

        # Log the entire API response
        logging.debug(f"API response for {appid}: {data}")

        if str(appid) in data and data[str(appid)]['success']:
            data = data[str(appid)]['data']

            # Apparently it had this field this whole time
            app_type = data.get('type', '').lower()

            logging.debug(f"AppID {appid} is of type {app_type}")

            if app_type == 'game':
                logging.debug(f"AppID {appid} is a game.")
                return data, False  # Return the data and False indicating it's a game

            else:
                logging.info(f"AppID {appid} is a not a game. Skipping.")
                return None, True  # Return the data and True indicating it's not a game

        logging.warning(f"Response for AppID {appid} marked as unsuccessful.")
        return None, False

    except requests.HTTPError as http_err:
        if response.status_code == 429:
            retry_after = response.headers.get('Retry-After')
            logging.error(
                f"HTTP 429 Rate Limit hit. Headers: {response.headers}")

            if retry_after:  # It never has this header but its supposed to so its here from my pain and suffering
                wait_time = int(retry_after)

                logging.info(
                    f"Rate limit hit for game details API. Waiting for {wait_time} seconds before retrying.")

                time.sleep(wait_time)
                return get_game_details(appid)
            else:
                wait_time = 60

                logging.warning(
                    "Rate limit hit for game details API, but no Retry-After header found. Waiting for 60 seconds before retrying.")

                time.sleep(wait_time)
                return None, False
        else:
            logging.error(
                f"HTTP error {response.status_code} occurred while fetching details for AppID {appid}: {http_err}")
            logging.error(f"Response: {response.text}")
            return None, False

    except requests.RequestException as req_err:
        logging.warning(
            f"Request error occurred while fetching details for AppID {appid}: {req_err}")
        return None, False

    except ValueError as json_err:
        logging.critical(
            f"JSON decoding failed for details of AppID {appid}: {json_err}. Response: {response.text}")
        return None, False


def get_game_reviews_summary(appid):
    url = f"https://store.steampowered.com/appreviews/{appid}?json=1"

    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()

    except requests.HTTPError as http_err:

        if response.status_code == 429:
            retry_after = response.headers.get('Retry-After')

            if retry_after:  # It actually never has this header, but it's supposed to so I'm leaving it in for fun.
                wait_time = int(retry_after)

                logging.info(
                    f"Rate limit hit for reviews API. Waiting for {wait_time} seconds before retrying.")

                time.sleep(wait_time)
                return get_game_reviews_summary(appid)

            else:
                wait_time = 60

                logging.warning(
                    "Rate limit hit for reviews API, but no Retry-After header found. Waiting for 60 seconds before retrying.")

                time.sleep(wait_time)
                return get_game_reviews_summary(appid)
        else:
            logging.error(
                f"HTTP error {response.status_code} occurred while fetching reviews for AppID {appid}: {http_err}")
            logging.error(f"Response: {response.text}")

    except requests.RequestException as req_err:
        logging.warning(
            f"Request error occurred while fetching reviews for AppID {appid}: {req_err}")

    except ValueError as json_err:
        logging.critical(
            f"JSON decoding failed for reviews of AppID {appid}: {json_err}. Response: {response.text}")

    return None


def format_timestamp(timestamp):
    return datetime.utcfromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')


def main():
    unwanted_apps = []
    stop_event = threading.Event()
    save_thread = threading.Thread(target=save_to_db_interval, args=(
        300, stop_event))  # Save every 5 minutes (300 seconds)

    try:
        # start saving thread
        save_thread.start()

        # Register the signal handler for Ctrl+C
        signal.signal(signal.SIGINT, signal_handler)

        load_unwanted_games()

        # Check to see if there is a list of remaning appids to process
        remaining_games = load_remaning_app_list()

        if remaining_games is not None:
            games = remaining_games
            games = [game for game in remaining_games  # basically get rid of the games we don't want from the saved json and the word list
                     # so we can update the word list and restart for future games
                     if game["appid"] not in unwanted_apps
                     and not any(keyword.lower() in game["name"].lower() for keyword in excluded_keywords)]

            logging.info("Resuming from saved AppID list.")
            logging.info(f"Total games for processing: {len(games)}")

        else:
            games = get_steam_games()  # otherwise get the list of games from the api
            # the function already checks for the word list
            games = [game for game in games if game["appid"]
                     not in unwanted_apps]

            logging.warning(
                "No saved AppID list found. Fetching list of games from Steam API.")
            logging.info(f"Total games for processing: {len(games)}")

        # i is used for saving the remaining gameids for resuming the script
        for i in range(len(games)):
            game = games[i]
            # rate limit is 200 requests per 5 minutes, so 1 request every 1.5 seconds should be fine
            time.sleep(1.5)

            # Check for Ctrl+C to stop the script and save data
            if interrupted:
                remaining_games = [{"appid": game["appid"],
                                    "name": game["name"]} for game in games[i:]]
                stop_event.set()
                save_thread.join()
                save_remaning_apps_list(remaining_games)
                save_unwanted_apps_list(unwanted_apps)
                write_data_to_sqlite(game_data)
                export_db_to_csv()
                logging.info(
                    "Data and AppID list saved due to signal interrupt. Exiting the script.")
                sys.exit(0)

            appid = game["appid"]

            logging.info(
                f"Fetching details for {game['name']} (AppID: {appid})...")
            details, is_not_game = get_game_details(appid)

            # wanted seperation logging to make sure it was working
            if details is None:
                logging.debug(
                    f"No valid details for {game['name']} (AppID: {appid}), skipping.")
                continue

            if is_not_game:
                unwanted_apps.append(appid)
                logging.debug(f"AppID {appid} is not a game, skipping.")
                continue

            reviews_summary = get_game_reviews_summary(appid)

            if reviews_summary is None:
                logging.warning(
                    f"No valid reviews summary for {game['name']} (AppID: {appid}), skipping.")
                continue

            if details:
                name = details.get('name', 'No Name Available')
                genres = ", ".join(
                    [genre['description'] for genre in details['genres']]) if 'genres' in details else 'N/A'
                # Extract the game description
                short_description = details.get('short_description', 'N/A')

                if 'query_summary' in reviews_summary and reviews_summary['query_summary']['total_reviews'] > 0:
                    qs = reviews_summary['query_summary']
                    positive_percentage = (
                        qs['total_positive'] / qs['total_reviews']) * 100 if qs['total_reviews'] > 0 else 0
                    review = reviews_summary['reviews'][0] if reviews_summary['reviews'] else None

                    try:
                        review_text = review['review'] if review else 'N/A'
                        # Check if review text is in English, replace line breaks
                        if is_english(review_text):
                            review_text = review_text.replace(
                                '\n', ' ').replace('\r', '')

                        else:
                            review_text = "N/A"

                    except Exception as e:
                        logging.warning(
                            f"Error processing review text for {name} (AppID: {appid}): {e}")
                        review_text = 'Error processing review text'

                    with data_lock:
                        game_data.append({
                            'Name': name,
                            'AppID': appid,
                            'Short Description': short_description,
                            'Review Score': qs['review_score'],
                            'Review Score Description': qs['review_score_desc'],
                            'Total Positive': qs['total_positive'],
                            'Total Negative': qs['total_negative'],
                            'Total Reviews': qs['total_reviews'],
                            'Positive Review Percentage': positive_percentage,
                            'Review Text (First Review)': review_text,
                            'Playtime Forever (First Review)': review['author'].get('playtime_forever', 'N/A') if review else 'N/A',
                            'Playtime Last Two Weeks (First Review)': review['author'].get('playtime_last_two_weeks', 'N/A') if review else 'N/A',
                            'Playtime at Review (First Review)': review['author'].get('playtime_at_review', 'N/A') if review else 'N/A',
                            'Timestamp Created (First Review)': format_timestamp(review['timestamp_created']) if review else 'N/A',
                            'Timestamp Updated (First Review)': format_timestamp(review['timestamp_updated']) if review else 'N/A',
                            'Voted Up (First Review)': review.get('voted_up', 'N/A') if review else 'N/A',
                            'Votes Up (First Review)': review.get('votes_up', 'N/A') if review else 'N/A',
                            'Genres': genres
                        })
                else:
                    logging.info(
                        f"No reviews for {name} (AppID: {appid}), skipping.")
                    unwanted_apps.append(appid)
            else:
                logging.warning(
                    f"Game details fetch unsuccessful for AppID: {appid}.")

    except Exception as e:
        logging.critical(f"An unexpected error occurred: {e}")
        logging.critical("Full traceback:")
        logging.critical(traceback.format_exc())
        logging.info("Saving the current data before exiting.")

    finally:
        # Save all the data if the script is interrupted
        if not interrupted:
            stop_event.set()
            save_thread.join()
            write_data_to_sqlite(game_data)
            save_unwanted_apps_list(unwanted_apps)
            export_db_to_csv()
            logging.critical(
                "Something went wrong, in finally block. Data saved")


if __name__ == "__main__":
    main()
